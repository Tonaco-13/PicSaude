"""
Gera a base de medicamentos do prescritor a partir da CMED/ANVISA (conformidade XLSX).

Modelo B (prescritível): deduplica as ~25k apresentações comerciais em unidades
`princípio ativo · forma · concentração`, agregando as marcas (PRODUTO) como aliases.
Saída principal: `data/def_medicamentos.csv` (esquema lido por app/ai/lookup_def.py).
Saída extra (Modelo A, completude): `data/cmed_apresentacoes.csv` (toda apresentação).

Fonte: ANVISA/CMED — Lista de Preços de Medicamentos (conformidade), coluna APRESENTAÇÃO
parseada por abreviações farmacêuticas padronizadas (com concordância de gênero, etc.).

Uso:
    python backend/scripts/gerar_medicamentos_cmed.py --xlsx CEMED/xls_conformidade_site_*.xlsx
"""
from __future__ import annotations

import argparse
import csv
import glob
import re
import unicodedata
from pathlib import Path

# NB: openpyxl é importado dentro de main() (lazy) — é dependência só do gerador
# offline, não do runtime; assim os testes do parser não exigem openpyxl no CI.

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_DEF = PROJECT_ROOT / "data" / "def_medicamentos.csv"
OUT_APRES = PROJECT_ROOT / "data" / "cmed_apresentacoes.csv"
_VERSAO = "ANVISA/CMED 2026-06"

COL = {"substancia": 0, "produto": 8, "apresentacao": 9, "classe": 10, "registro": 4, "tarja": 72}

# --- Formas farmacêuticas (token canônico + variantes de abreviação) -----------
_FORMA = {
    "COM": "comprimido", "COMP": "comprimido", "CPR": "comprimido",
    "CAP": "cápsula", "CAPS": "cápsula",
    "DRG": "drágea", "DRGE": "drágea",
    "SOL": "solução", "SUS": "suspensão", "SUSP": "suspensão",
    "EMU": "emulsão", "EMUL": "emulsão",
    "XPE": "xarope", "ELIX": "elixir", "ELX": "elixir",
    "PO": "pó", "PÓ": "pó", "GRAN": "granulado", "GRA": "granulado",
    "CREM": "creme", "CR": "creme", "POM": "pomada",
    "GEL": "gel", "PAST": "pasta", "PAS": "pastilha", "PASTIL": "pastilha",
    "SUP": "supositório", "OVU": "óvulo", "OVL": "óvulo",
    "ADES": "adesivo", "AER": "aerossol", "SPRAY": "spray", "GTS": "gotas",
    "LIQ": "líquido", "LOC": "loção", "SHA": "xampu", "SHAMP": "xampu", "XAMP": "xampu",
    "ESM": "esmalte", "IMPL": "implante", "PASTIL": "pastilha", "GOM": "goma", "GOMA": "goma",
    "FILM": "filme", "FIL": "filme", "COLUT": "colutório", "SAB": "sabonete", "GLB": "glóbulo",
}
# Atributos/qualificadores (revestido, liberação, etc.) e rotas como adjetivos (gênero masc. base).
_ATRIB = {
    "REV": "revestido", "REVEST": "revestido", "DURA": "dura", "MOLE": "mole",
    "GEL": "gelatinosa", "LIOF": "liofilizado", "ORODISP": "orodispersível",
    "SUBL": "sublingual", "MAST": "mastigável", "EFERV": "efervescente", "EFEV": "efervescente",
    "DISP": "dispersível", "RETARD": "retardada", "PROL": "prolongada", "LIB": "de liberação",
    "TRANSD": "transdérmico", "TRANS": "transdérmico", "INAL": "para inalação",
    "DOSIF": "dosimetrado", "IVIT": "intravítreo",
    "VAG": "vaginal", "VG": "vaginal",
    "DERM": "dermatológico", "OFT": "oftálmico", "NAS": "nasal", "OT": "otológico",
    "AURIC": "auricular", "RET": "retal", "TOP": "tópico", "CAPI": "capilar",
    "OR": "oral", "INJ": "injetável", "INFUS": "para infusão", "DIL": "para diluição",
}
_VIA = {
    "IV": "intravenosa", "IM": "intramuscular", "SC": "subcutânea", "ID": "intradérmica",
    "OR": "oral", "DERM": "tópica", "OFT": "oftálmica", "NAS": "nasal",
    "OT": "otológica", "AURIC": "auricular", "VAG": "vaginal", "RET": "retal",
    "TOP": "tópica", "CAPI": "capilar", "SL": "sublingual",
}
# Rotas parenterais "puras": vão para a VIA, nunca compõem o nome da forma.
_VIA_PURA = {"IV", "IM", "SC", "ID"}
_FORMA_STARTERS = set(_FORMA)
_EMBAL = {"CT", "FR", "EMB", "BG", "POT", "BISN", "CX", "BL", "ENV", "TB", "TUBO",
          "SER", "AMP", "FA", "SG", "SISTEMA", "CAIXA", "FRASCO", "SAC"}
# nomes de forma femininos (para concordância de gênero das rotas-adjetivo)
_FEM = {"cápsula", "solução", "suspensão", "emulsão", "pomada", "pasta", "loção",
        "goma", "pastilha", "gotas", "drágea"}
_ROTAS_GENERO = {"dermatológico", "oftálmico", "tópico", "otológico"}


def _strip_acentos(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _pre_norm(ap: str) -> str:
    s = (ap or "").upper().strip()
    s = re.sub(r"(?<=\d)(?=[A-ZÀ-Ÿ])", " ", s)   # "178MG" -> "178 MG"
    s = re.sub(r"(?<=[A-ZÀ-Ÿ])\.", "", s)         # "SOL." -> "SOL" (sem tocar números)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _limpar_substancia(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"^\d+\s*-\s*", "", s)
    s = s.replace(";", " + ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.lower()


def _concordar_genero(forma: str) -> str:
    palavras = forma.split()
    if not palavras:
        return forma
    fem = palavras[0] in _FEM
    out = []
    for w in palavras:
        if w in _ROTAS_GENERO and fem:
            out.append(w[:-1] + "a")   # dermatológico -> dermatológica
        else:
            out.append(w)
    return " ".join(out)


def _polir_forma(forma: str) -> str:
    forma = re.sub(r"\bcápsula gel\b", "cápsula gelatinosa", forma)
    # "pó/granulado [liofilizado] suspensão/solução" -> insere "para"
    forma = re.sub(r"\b(pó|granulado)( liofilizado)? (suspensão|solução)",
                   r"\1\2 para \3", forma)
    # "pó [liofilizado] injetável" -> "pó [liofilizado] para solução injetável" (canônico ANVISA)
    forma = re.sub(r"\bpó( liofilizado)? injetável\b", r"pó\1 para solução injetável", forma)
    # remove palavras consecutivas repetidas (ex.: 'transdérmico transdérmico' de TRANSD+TRANS)
    forma = re.sub(r"\b(\w+)( \1\b)+", r"\1", forma)
    # 'inalação nasal' (ex.: anestésico volátil com NAS espúrio) -> só 'inalação'
    forma = re.sub(r"\binalação nasal\b", "inalação", forma)
    forma = _concordar_genero(forma)
    return re.sub(r"\s+", " ", forma).strip()


def _parse_apresentacao(ap: str) -> tuple[str, str, str]:
    ap = _pre_norm(ap)
    if not ap:
        return "", "", ""
    toks = ap.split()
    idx_forma = next((i for i, t in enumerate(toks) if t in _FORMA_STARTERS), None)
    if idx_forma is None:
        return ap.lower(), "", ""

    dose = " ".join(toks[:idx_forma]).replace("(", "").replace(")", "")
    dose = re.sub(r"\s+", " ", dose).strip().lower()

    forma_toks, vias = [], []
    j = idx_forma
    while j < len(toks) and toks[j] not in _EMBAL:
        t = toks[j]
        if "/" in t and all(p in _VIA for p in t.split("/")):
            vias.extend(t.split("/"))
        elif t in _VIA:
            vias.append(t)
        forma_toks.append(t)
        j += 1
        if len(forma_toks) > 8:
            break

    legivel = []
    for t in forma_toks:
        if t in _VIA_PURA or "/" in t:
            continue                       # rotas parenterais e siglas compostas vão só p/ via
        if t in _FORMA:
            legivel.append(_FORMA[t])
        elif t in _ATRIB:
            legivel.append(_ATRIB[t])
        else:
            legivel.append(t.lower())
    forma = _polir_forma(" ".join(legivel))

    vset = [_VIA[v] for v in dict.fromkeys(vias) if v in _VIA]
    if "inala" in forma:           # inalatório (ex.: anestésico volátil) tem prioridade sobre NAS
        via = "inalatória"
    elif vset:
        via = ", ".join(vset)
    else:
        via = _inferir_via(forma)
    return dose, forma, via


def _inferir_via(forma: str) -> str:
    """Infere a via de administração a partir da forma quando não há sigla explícita."""
    if not forma:
        return ""
    head = forma.split()[0]
    if "intravítreo" in forma:
        return "intravítrea"
    if "transdérmic" in forma or head == "adesivo":
        return "transdérmica"
    if "inala" in forma or ("dosimetrado" in forma and "nasal" not in forma):
        return "inalatória"
    if "oftálmic" in forma:
        return "oftálmica"
    if "otológic" in forma or "auricular" in forma:
        return "otológica"
    if "vaginal" in forma or head == "óvulo":
        return "vaginal"
    if "nasal" in forma:
        return "nasal"
    if "retal" in forma or head == "supositório":
        return "retal"
    if head in ("pastilha", "goma", "colutório") or "sublingual" in forma:
        return "bucal"
    if (head in ("loção", "sabonete", "xampu", "esmalte", "creme", "pomada", "pasta", "gel")
            or "dermatológic" in forma or "tópic" in forma or "capilar" in forma):
        return "tópica"
    if (head in ("comprimido", "cápsula", "drágea", "xarope", "elixir", "granulado")
            or "oral" in forma):
        return "oral"
    if "injetável" in forma or "infusão" in forma:
        return "parenteral"
    return ""


def _tarja_controle(tarja: str) -> str:
    t = _strip_acentos((tarja or "").lower())
    if "preta" in t:
        return "controle_especial_notificacao"
    if "sob restricao" in t:
        return "venda_sob_prescricao_retida"
    if "vermelha" in t:
        return "venda_sob_prescricao"
    if "sem tarja" in t:
        return "venda_livre"
    return "nao_classificado"


def main() -> None:
    import openpyxl  # lazy — só o gerador offline precisa
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()
    caminho = sorted(glob.glob(args.xlsx))[0]
    print(f"Lendo {caminho} ...")

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    hdr = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        celulas = [_strip_acentos(str(c or "")).upper().strip() for c in row]
        if any(c == "SUBSTANCIA" for c in celulas) and any("APRESENTAC" in c for c in celulas):
            hdr = i
            break
    if hdr is None:
        raise SystemExit("cabeçalho da CMED não encontrado")

    apres_rows, grupos = [], {}
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        substancia_raw = (row[COL["substancia"]] or "").strip() if row[COL["substancia"]] else ""
        if not substancia_raw:
            continue
        produto = (row[COL["produto"]] or "").strip()
        apresentacao = (row[COL["apresentacao"]] or "").strip()
        classe = (row[COL["classe"]] or "").strip()
        tarja = (row[COL["tarja"]] or "").strip() if len(row) > COL["tarja"] else ""

        principio = _limpar_substancia(substancia_raw)
        conc, forma, via = _parse_apresentacao(apresentacao)
        controle = _tarja_controle(tarja)

        apres_rows.append({
            "principio_ativo": principio, "produto": produto, "apresentacao": apresentacao,
            "forma_farmaceutica": forma, "concentracao_texto": conc, "via_administracao": via,
            "classe_terapeutica": classe, "controle": controle, "fonte": _VERSAO,
        })
        chave = (principio, forma, conc)
        g = grupos.setdefault(chave, {
            "via_administracao": via, "controle": controle, "marcas": set(),
        })
        if produto:
            g["marcas"].add(produto.strip())

    OUT_DEF.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_DEF, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["principio_ativo", "nome_normalizado", "forma_farmaceutica",
                    "unidade_dispensavel", "concentracao_texto", "via_administracao",
                    "aliases", "fonte", "versao_base"])
        for (principio, forma, conc), g in sorted(grupos.items()):
            nome = " ".join(p for p in (principio, conc, forma) if p)
            unidade = forma.split()[0] if forma else ""
            aliases = "|".join(sorted(m for m in g["marcas"] if m)[:8])
            w.writerow([principio, nome, forma, unidade, conc, g["via_administracao"],
                        aliases, _VERSAO, "2026-06"])

    with open(OUT_APRES, "w", encoding="utf-8", newline="") as f:
        cols = ["principio_ativo", "produto", "apresentacao", "forma_farmaceutica",
                "concentracao_texto", "via_administracao", "classe_terapeutica", "controle", "fonte"]
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(apres_rows)

    vazias = sum(1 for r in apres_rows if not r["forma_farmaceutica"])
    print(f"OK — Modelo B: {len(grupos):,} unidades prescritíveis em {OUT_DEF}")
    print(f"     Modelo A: {len(apres_rows):,} apresentações em {OUT_APRES}")
    print(f"     forma vazia (não parseada): {vazias} ({vazias*100//max(len(apres_rows),1)}%)")


if __name__ == "__main__":
    main()
